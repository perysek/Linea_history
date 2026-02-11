"""
Migration script to import data from Excel file PPM_wew.xlsm
into the database with MOSYS data enrichment.

Excel file location: G:\DOCUMENT\qualita\System Zarządzania Jakością\Cele jakościowe\PPM_wew.xlsm

Features:
- Automatic duplicate detection based on (data_niezgodnosci, nr_raportu)
- Batch MOSYS data fetching for performance
- Defect parsing from Uwagi column
- Date-based filtering to skip old records
- Support for .xlsm (macro-enabled) and .xlsx files
- Verbose row-by-row status display

Usage:
    python migrate_excel_data.py                         # Normal import (verbose by default)
    python migrate_excel_data.py --dry-run               # Preview without committing
    python migrate_excel_data.py --quiet                 # Disable verbose row-by-row output
    python migrate_excel_data.py --start-row=10291       # Custom starting row
    python migrate_excel_data.py --from-date=2026-01-01  # Only import from this date
    python migrate_excel_data.py --from-date=2026-01-01 --dry-run  # Combine filters

Troubleshooting:
    If you see mostly empty cells or None values:
    1. Open the Excel file
    2. Select all data (Ctrl+A)
    3. Copy (Ctrl+C)
    4. Paste values only (Ctrl+Alt+V, then V)
    5. Save and try again

    This converts formulas to their cached values.
"""

import openpyxl
import pandas as pd
from datetime import datetime
import time
import os
from app import create_app, db
from app.models.sorting_area import DaneRaportu, BrakiDefektyRaportu, Operator
from MOSYS_data_functions import get_batch_niezgodnosc_details
import re


def parse_defects_from_uwagi(uwagi_text):
    """
    Parse defects from the Uwagi column.
    Example: "działania korygujące na pęcherze x52, nadpalenia x0"
    Returns list of tuples: [(defect_name, count), ...]
    """
    if not uwagi_text:
        return []

    defects = []
    # Pattern to match defects like "pęcherze x52" or "nadpalenia x0"
    pattern = r'([a-zA-ZąćęłńóśźżĄĆĘŁŃÓŚŹŻ\s]+)\s*x\s*(\d+)'
    matches = re.findall(pattern, uwagi_text)

    for defect_name, count in matches:
        defect_name = defect_name.strip()
        count = int(count)
        if count > 0:  # Only add defects with count > 0
            defects.append((defect_name, count))

    return defects


def convert_to_boolean(value):
    """Convert Excel value to boolean for 'Selekcja na bieżąco' field."""
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().lower() in ('x', 'tak', 'yes', 'true', '1')
    return bool(value)


def import_data_from_excel(excel_file=r'G:\DOCUMENT\qualita\System Zarządzania Jakością\Cele jakościowe\PPM wewnętrzny koszty złej jakości (2023).xlsm',
                           sheet_name='dane', start_row=2, dry_run=False,
                           batch_size=100, from_date=None, use_pandas=True, verbose=True):
    """
    Import data from Excel file into the database with MOSYS enrichment.

    Automatically skips duplicate records based on (data_niezgodnosci, nr_raportu) combination.

    Args:
        excel_file: Path to the Excel file
        sheet_name: Name of the sheet to read
        start_row: Excel row number to start from (1-indexed, includes header at row 1)
        dry_run: If True, don't commit to database, just print what would be imported
        batch_size: Number of records to process in each batch
        from_date: If provided, only import records with data_niezgodnosci >= this date (date object or None)
        use_pandas: If True, use pandas to read Excel (much faster for large files). Default: True
        verbose: If True, show detailed status for each row processed. Default: True
    """
    print(f"\n[DEBUG] Inside import_data_from_excel()")
    print(f"[DEBUG] Excel file: {excel_file}")
    print(f"[DEBUG] Sheet name: {sheet_name}")
    print(f"[DEBUG] Start row: {start_row}")

    print(f"\n[DEBUG] Creating Flask app...")
    try:
        app = create_app()
        print(f"[DEBUG] [OK] Flask app created")
    except Exception as e:
        print(f"[ERROR] Failed to create Flask app: {e}")
        raise

    start_time = time.time()

    print(f"[DEBUG] Entering app context...")
    with app.app_context():
        # Validate file exists
        if not os.path.exists(excel_file):
            print(f"ERROR: Excel file not found: {excel_file}")
            print(f"Current directory: {os.getcwd()}")
            print(f"Please ensure the file exists in the correct location.")
            return

        print(f"\n[DEBUG] Loading Excel file: {excel_file}")
        file_size = os.path.getsize(excel_file) / (1024 * 1024)  # Size in MB
        print(f"[DEBUG] File size: {file_size:.2f} MB")

        # Load Excel data
        print(f"[DEBUG] Starting to load Excel data...")
        if use_pandas:
            # Use pandas for faster loading (especially for large .xlsm files)
            print(f"[DEBUG] Using pandas for fast Excel reading...")
            try:
                # Read without treating first row as header to preserve all rows
                df = pd.read_excel(excel_file, sheet_name=sheet_name, engine='openpyxl', header=None)
                print(f"[DEBUG] [OK] Excel data loaded successfully with pandas")
                print(f"[DEBUG] Total rows: {len(df)}, Total columns: {len(df.columns)}")

                # Convert DataFrame to list of lists (mimicking openpyxl structure)
                # Add None at index 0 so that excel_data[1] = Excel row 1, excel_data[2] = Excel row 2, etc.
                excel_data = [None] + df.values.tolist()
                ws_max_row = len(excel_data) - 1  # Subtract 1 for the dummy element

                print(f"[DEBUG] Data converted to internal format")
                print(f"[DEBUG] Excel list length: {len(excel_data)}, accessible rows: 1-{ws_max_row}")

            except Exception as e:
                print(f"[ERROR] Failed loading Excel with pandas: {e}")
                print(f"Falling back to openpyxl (this will be slower)...")
                import traceback
                traceback.print_exc()
                use_pandas = False

        if not use_pandas:
            # Fallback to openpyxl
            print(f"[DEBUG] Using openpyxl to load Excel...")
            try:
                if excel_file.endswith('.xlsm'):
                    print("[DEBUG] Detected macro-enabled workbook (.xlsm)")
                    print("[DEBUG] WARNING: This may take 5-10 minutes for large files...")
                    wb = openpyxl.load_workbook(excel_file, data_only=True, keep_vba=False)
                else:
                    wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
                print(f"[DEBUG] [OK] Workbook loaded")

                ws = wb[sheet_name]
                print(f"[DEBUG] [OK] Worksheet '{sheet_name}' found")
                ws_max_row = ws.max_row
                excel_data = None  # Will read rows on demand

            except Exception as e:
                print(f"[ERROR] Failed loading workbook: {e}")
                import traceback
                traceback.print_exc()
                return

        print(f"\n[DEBUG] Total rows in Excel: {ws_max_row}")
        print(f"Starting from row: {start_row}")
        if from_date:
            print(f"Filtering from date: {from_date}")
        if dry_run:
            print("=== DRY RUN MODE - No data will be committed ===")

        # Quick validation - check if we can read the first data row
        print(f"\nValidating data access...")
        try:
            if use_pandas:
                test_row = excel_data[start_row]
            else:
                test_row = [cell.value for cell in ws[start_row]]
            non_empty = sum(1 for cell in test_row if cell is not None)
            print(f"[OK] First row has {non_empty} non-empty cells")
            if non_empty < 5:
                print(f"WARNING: First row seems mostly empty.")
        except Exception as e:
            print(f"WARNING: Could not validate first row: {e}")
        print()

        # Fixed values for all records
        OPERATOR_ID = 2
        NR_INSTRUKCJI = 'wg raportu'

        # Verify operator exists
        print(f"\n[DEBUG] Checking database connection and operator...")
        try:
            operator = db.session.query(Operator).filter_by(id=OPERATOR_ID).first()
            print(f"[DEBUG] [OK] Database query successful")
        except Exception as e:
            print(f"[ERROR] Database query failed: {e}")
            import traceback
            traceback.print_exc()
            return
        if not operator:
            print(f"WARNING: Operator with ID {OPERATOR_ID} not found in database!")
            print("Migration will continue but operator_id will be set to {OPERATOR_ID}.")
            print("Make sure this operator exists or update OPERATOR_ID in the script.\n")

        imported_count = 0
        skipped_count = 0
        skipped_by_date = 0
        duplicate_count = 0
        error_count = 0
        mosys_fetch_count = 0

        # Column mapping based on PPM_import_excel.csv (Excel columns are 1-indexed)
        COL_LP = 1  # nr_raportu
        COL_DATA_NIEZGODNOSCI = 2  # Will be overridden by MOSYS data
        COL_NR_NIEZGODNOSCI = 5
        COL_SELEKCJA_NA_BIEZACO = 7
        COL_ILOSC_SPRAWDZONYCH = 9
        COL_ILOSC_WADLIWYCH = 11
        COL_ZALECANA_WYDAJNOSC = 13
        COL_CZAS_PRACY = 19
        COL_UWAGI = 21
        COL_UWAGI_DO_WYDAJNOSCI = 22
        COL_DATA_SELEKCJI = 26

        # Process records in batches for MOSYS efficiency
        current_row = start_row
        total_rows = ws_max_row - start_row + 1
        batch_num = 0

        while current_row <= ws_max_row:
            batch_end = min(current_row + batch_size, ws_max_row + 1)
            batch_num += 1

            # Calculate progress
            rows_processed = current_row - start_row
            progress_pct = (rows_processed / total_rows * 100) if total_rows > 0 else 0

            print(f"\n{'='*60}")
            print(f"Batch {batch_num} | Rows {current_row}-{batch_end-1} | Progress: {progress_pct:.1f}%")
            print(f"{'='*60}")

            # Step 1: Collect batch data from Excel
            print(f"[1/3] Reading Excel rows...")
            if verbose:
                print(f"\n{'Row':<8} {'NC Number':<15} {'Nr Rap':<8} {'Qty Sorted':<12} {'Qty Scrapped':<13} {'Status':<20}")
                print(f"{'-'*8} {'-'*15} {'-'*8} {'-'*12} {'-'*13} {'-'*20}")

            batch_data = []
            nr_niezgodnosci_list = []

            for row_num in range(current_row, batch_end):
                try:
                    # Get row data (pandas or openpyxl)
                    if use_pandas:
                        if row_num >= len(excel_data):
                            # Row doesn't exist (beyond actual data)
                            skipped_count += 1
                            if verbose:
                                # Debug: show why row is skipped
                                reason = f"row {row_num} >= len {len(excel_data)}"
                                print(f"{row_num:<8} {'---':<15} {'---':<8} {'---':<12} {'---':<13} {'SKIP: ' + reason:<20}")
                            continue
                        row = excel_data[row_num]
                        # Convert pandas NaN to None for SQL compatibility
                        row = [None if (isinstance(cell, float) and pd.isna(cell)) else cell for cell in row]
                    else:
                        row = [cell.value for cell in ws[row_num]]

                    # Skip if nr_niezgodnosci is empty
                    nr_niezgodnosci = row[COL_NR_NIEZGODNOSCI - 1]
                    if not nr_niezgodnosci:
                        skipped_count += 1
                        if verbose:
                            nr_rap = row[COL_LP - 1] if row[COL_LP - 1] else '---'
                            print(f"{row_num:<8} {'<empty>':<15} {str(nr_rap):<8} {'---':<12} {'---':<13} {'SKIPPED: No NC':<20}")
                        continue

                    nr_niezgodnosci = str(nr_niezgodnosci).strip()
                    nr_niezgodnosci_list.append(nr_niezgodnosci)

                    # Extract Excel data for this row
                    row_data = {
                        'row_num': row_num,
                        'nr_raportu': str(row[COL_LP - 1]) if row[COL_LP - 1] is not None else None,
                        'nr_niezgodnosci': nr_niezgodnosci,
                        'selekcja_na_biezaco': convert_to_boolean(row[COL_SELEKCJA_NA_BIEZACO - 1]),
                        'ilosc_detali_sprawdzonych': row[COL_ILOSC_SPRAWDZONYCH - 1],
                        'ilosc_wadliwych': row[COL_ILOSC_WADLIWYCH - 1],
                        'zalecana_wydajnosc': row[COL_ZALECANA_WYDAJNOSC - 1],
                        'czas_pracy': row[COL_CZAS_PRACY - 1],
                        'uwagi': row[COL_UWAGI - 1],
                        'uwagi_do_wydajnosci': row[COL_UWAGI_DO_WYDAJNOSCI - 1],
                        'data_selekcji': row[COL_DATA_SELEKCJI - 1],
                    }

                    # Convert datetime to date for data_selekcji
                    if isinstance(row_data['data_selekcji'], datetime):
                        row_data['data_selekcji'] = row_data['data_selekcji'].date()

                    # Filter by date if from_date is specified
                    if from_date:
                        excel_date_niezgodnosci = row[COL_DATA_NIEZGODNOSCI - 1]
                        # Convert to date if it's a datetime object
                        if isinstance(excel_date_niezgodnosci, datetime):
                            excel_date_niezgodnosci = excel_date_niezgodnosci.date()

                        # Skip if date is before from_date
                        if excel_date_niezgodnosci and excel_date_niezgodnosci < from_date:
                            skipped_by_date += 1
                            if verbose:
                                print(f"{row_num:<8} {str(nr_niezgodnosci):<15} {str(row_data['nr_raportu']):<8} {'---':<12} {'---':<13} {'FILTERED: < ' + str(from_date):<20}")
                            continue

                    # Show row will be processed (before MOSYS fetch)
                    if verbose:
                        qty_sorted = row_data['ilosc_detali_sprawdzonych'] if row_data['ilosc_detali_sprawdzonych'] else 0
                        qty_scrapped = row_data['ilosc_wadliwych'] if row_data['ilosc_wadliwych'] else 0
                        print(f"{row_num:<8} {str(nr_niezgodnosci):<15} {str(row_data['nr_raportu']):<8} {str(qty_sorted):<12} {str(qty_scrapped):<13} {'Fetching MOSYS...':<20}", end='\r')

                    batch_data.append(row_data)

                except Exception as e:
                    error_count += 1
                    print(f"Error reading Excel row {row_num}: {e}")
                    continue

            if not batch_data:
                print(f"  -> No valid data in this batch, skipping...")
                current_row = batch_end
                continue

            print(f"  -> Collected {len(batch_data)} valid rows from Excel")

            # Step 2: Batch fetch MOSYS data for all nr_niezgodnosci in batch
            print(f"[2/3] Fetching MOSYS data for {len(nr_niezgodnosci_list)} NC numbers...")
            try:
                mosys_data = get_batch_niezgodnosc_details(nr_niezgodnosci_list)
                mosys_fetch_count += len(nr_niezgodnosci_list)
                print(f"  -> Successfully fetched MOSYS data for {len(mosys_data)} records")
            except Exception as e:
                print(f"  [X] Error fetching MOSYS data: {e}")
                mosys_data = {}

            # Step 3: Create database records for batch
            print(f"\n[3/3] Creating database records...")
            if verbose:
                print(f"\n{'Row':<8} {'NC Number':<15} {'Nr Rap':<8} {'MOSYS Date':<12} {'Order':<10} {'Part':<12} {'Status':<20}")
                print(f"{'-'*8} {'-'*15} {'-'*8} {'-'*12} {'-'*10} {'-'*12} {'-'*20}")

            batch_imported_before = imported_count
            batch_duplicates_before = duplicate_count

            for record in batch_data:
                try:
                    nr_niezg = record['nr_niezgodnosci']

                    # Get MOSYS data for this record
                    mosys_info = mosys_data.get(nr_niezg, {})
                    data_niezgodnosci = mosys_info.get('data_niezgodnosci')
                    nr_zamowienia = mosys_info.get('nr_zamowienia')
                    kod_detalu = mosys_info.get('kod_detalu')

                    # Check for duplicate based on data_niezgodnosci + nr_raportu
                    existing_record = db.session.query(DaneRaportu).filter_by(
                        data_niezgodnosci=data_niezgodnosci,
                        nr_raportu=record['nr_raportu']
                    ).first()

                    if existing_record:
                        duplicate_count += 1
                        if verbose:
                            print(f"{record['row_num']:<8} {str(nr_niezg):<15} {str(record['nr_raportu']):<8} "
                                  f"{str(data_niezgodnosci) if data_niezgodnosci else 'N/A':<12} "
                                  f"{str(nr_zamowienia)[:10] if nr_zamowienia else 'N/A':<10} "
                                  f"{str(kod_detalu)[:12] if kod_detalu else 'N/A':<12} "
                                  f"{'DUPLICATE':<20}")
                        elif dry_run and duplicate_count <= 3:
                            print(f"  [DUPLICATE] Skipping row {record['row_num']}: "
                                  f"nr_raportu={record['nr_raportu']}, "
                                  f"data_niezgodnosci={data_niezgodnosci}")
                        continue

                    # Create DaneRaportu record
                    raport = DaneRaportu(
                        nr_raportu=record['nr_raportu'],
                        operator_id=OPERATOR_ID,
                        nr_niezgodnosci=nr_niezg,
                        data_niezgodnosci=data_niezgodnosci,  # From MOSYS
                        nr_zamowienia=nr_zamowienia,  # From MOSYS
                        kod_detalu=kod_detalu,  # From MOSYS
                        nr_instrukcji=NR_INSTRUKCJI,  # Fixed value
                        selekcja_na_biezaco=record['selekcja_na_biezaco'],
                        ilosc_detali_sprawdzonych=record['ilosc_detali_sprawdzonych'],
                        zalecana_wydajnosc=record['zalecana_wydajnosc'],
                        czas_pracy=record['czas_pracy'],
                        uwagi=record['uwagi'],
                        uwagi_do_wydajnosci=record['uwagi_do_wydajnosci'],
                        data_selekcji=record['data_selekcji']
                    )

                    db.session.add(raport)
                    db.session.flush()  # Get the ID for the raport

                    # Print sample in dry run mode (first 5 records only)
                    if dry_run and imported_count < 5:
                        print(f"\n[Sample {imported_count + 1}] Would import:")
                        print(f"  Excel row: {record['row_num']}")
                        print(f"  Nr raportu: {record['nr_raportu']}")
                        print(f"  Nr niezgodnosci: {nr_niezg}")
                        print(f"  Data niezgodnosci: {data_niezgodnosci} (from MOSYS)")
                        print(f"  Nr zamowienia: {nr_zamowienia} (from MOSYS)")
                        print(f"  Kod detalu: {kod_detalu} (from MOSYS)")
                        print(f"  Data selekcji: {record['data_selekcji']}")
                        print(f"  Operator ID: {OPERATOR_ID}")
                        print(f"  Nr instrukcji: {NR_INSTRUKCJI}")

                    # Parse and add defects from Uwagi column
                    defects = parse_defects_from_uwagi(record['uwagi'])
                    for defect_name, count in defects:
                        defekt = BrakiDefektyRaportu(
                            raport_id=raport.id,
                            defekt=defect_name,
                            ilosc=count
                        )
                        db.session.add(defekt)

                    # If no specific defects parsed, use total count from column 11
                    ilosc_wadliwych = record['ilosc_wadliwych']
                    if ilosc_wadliwych and ilosc_wadliwych > 0 and not defects:
                        defekt = BrakiDefektyRaportu(
                            raport_id=raport.id,
                            defekt=record['uwagi'] if record['uwagi'] else "Niespecyfikowane",
                            ilosc=ilosc_wadliwych
                        )
                        db.session.add(defekt)

                    imported_count += 1

                    # Show success status
                    if verbose:
                        defect_count = len(defects) if defects else (1 if ilosc_wadliwych and ilosc_wadliwych > 0 else 0)
                        status = f"IMPORTED ({defect_count} defects)" if dry_run else f"IMPORTED ({defect_count} defects)"
                        print(f"{record['row_num']:<8} {str(nr_niezg):<15} {str(record['nr_raportu']):<8} "
                              f"{str(data_niezgodnosci) if data_niezgodnosci else 'N/A':<12} "
                              f"{str(nr_zamowienia)[:10] if nr_zamowienia else 'N/A':<10} "
                              f"{str(kod_detalu)[:12] if kod_detalu else 'N/A':<12} "
                              f"{status:<20}")

                except Exception as e:
                    error_count += 1
                    if verbose:
                        print(f"{record.get('row_num', '?'):<8} {str(record.get('nr_niezgodnosci', '?')):<15} "
                              f"{str(record.get('nr_raportu', '?')):<8} {'---':<12} {'---':<10} {'---':<12} "
                              f"{'ERROR: ' + str(e)[:20]:<20}")
                    else:
                        print(f"Error processing row {record.get('row_num', '?')}: {e}")
                    db.session.rollback()
                    continue

            # Commit batch (or rollback if dry run)
            batch_imported = imported_count - batch_imported_before
            batch_duplicates = duplicate_count - batch_duplicates_before

            if dry_run:
                db.session.rollback()
                print(f"\n[OK] [DRY RUN] Batch complete:")
                print(f"  -> {batch_imported} new records (would be imported)")
                print(f"  -> {batch_duplicates} duplicates skipped")
                print(f"  -> Total progress: {imported_count} imported, {duplicate_count} duplicates, {skipped_by_date} filtered by date")
            else:
                db.session.commit()
                print(f"\n[OK] Batch committed:")
                print(f"  -> {batch_imported} new records imported")
                print(f"  -> {batch_duplicates} duplicates skipped")
                print(f"  -> Total progress: {imported_count} imported, {duplicate_count} duplicates, {skipped_by_date} filtered by date")

            current_row = batch_end

        # Final summary
        elapsed_time = time.time() - start_time
        minutes, seconds = divmod(int(elapsed_time), 60)

        print(f"\n{'='*60}")
        if dry_run:
            print("[DRY RUN] Rolled back all changes - no data was committed")
        else:
            print("[OK] All changes committed to database")
        print(f"{'='*60}")

        print(f"\n=== Import Summary ===")
        print(f"Total rows processed: {ws_max_row - start_row + 1}")
        print(f"Successfully imported: {imported_count}")
        print(f"Skipped (no nr_niezgodnosci): {skipped_count}")
        if from_date:
            print(f"Skipped (before {from_date}): {skipped_by_date}")
        print(f"Duplicates skipped: {duplicate_count}")
        print(f"Errors: {error_count}")
        print(f"MOSYS lookups performed: {mosys_fetch_count}")
        print(f"\nTime elapsed: {minutes}m {seconds}s")


if __name__ == "__main__":
    import sys
    from datetime import date

    print("\n" + "="*60)
    print("STARTING Excel Data Migration Script")
    print("="*60)

    # Check for command-line arguments
    print("\n[DEBUG] Parsing command-line arguments...")
    dry_run = '--dry-run' in sys.argv or '-d' in sys.argv
    verbose = '--quiet' not in sys.argv  # Verbose by default, use --quiet to disable
    start_row = 1  # Default start row (based on database verification)
    from_date = None  # Default no date filtering

    # Allow custom start row and from_date from command line
    for arg in sys.argv:
        if arg.startswith('--start-row='):
            start_row = int(arg.split('=')[1])
        elif arg.startswith('--from-date='):
            date_str = arg.split('=')[1]
            try:
                # Parse date in format YYYY-MM-DD
                year, month, day = map(int, date_str.split('-'))
                from_date = date(year, month, day)
            except (ValueError, AttributeError) as e:
                print(f"Error parsing --from-date: {e}")
                print("Expected format: --from-date=YYYY-MM-DD (e.g., --from-date=2026-01-01)")
                sys.exit(1)

    print("\nExcel Data Migration Script (with MOSYS enrichment)")
    print("=" * 60)
    print(f"Start row: {start_row}")
    if from_date:
        print(f"From date: {from_date} (only records >= this date)")
    print(f"Dry run: {dry_run}")
    print(f"Verbose: {verbose} (use --quiet to disable)")
    print(f"Operator ID: 2 (fixed)")
    print(f"Nr instrukcji: 'wg raportu' (fixed)")
    print(f"MOSYS data: data_niezgodnosci, nr_zamowienia, kod_detalu")
    print("=" * 60)

    # Excel file location
    excel_file = r'G:\DOCUMENT\qualita\System Zarządzania Jakością\Cele jakościowe\PPM wewnętrzny koszty złej jakości (2023).xlsm'

    print(f"\n[DEBUG] Excel file location: {excel_file}")
    print("\n[DEBUG] Calling import_data_from_excel()...")
    try:
        import_data_from_excel(
            excel_file=excel_file,
            sheet_name='dane',
            start_row=start_row,
            dry_run=dry_run,
            from_date=from_date,
            use_pandas=True,  # Use pandas for fast loading (especially .xlsm files)
            verbose=verbose   # Show detailed row-by-row status
        )
        print("\n[DEBUG] import_data_from_excel() completed successfully")
    except Exception as e:
        print(f"\n[ERROR] Script failed with exception:")
        print(f"Error type: {type(e).__name__}")
        print(f"Error message: {e}")
        import traceback
        print("\nFull traceback:")
        traceback.print_exc()
        sys.exit(1)
